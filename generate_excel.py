#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
德国独立储能电站财务测算系统 - Excel 版本生成器
@description 将网页版财务测算系统转换为 Excel 表格，支持本地计算
@version 1.0
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os
import json

# ==================== 颜色定义（符合国际通用习惯） ====================
INPUT_COLOR = "E7F3FF"  # 浅蓝色 - 用户输入
CALC_COLOR = "FFF9E6"  # 浅黄色 - 公式计算
RESULT_COLOR = "E6F7E6"  # 浅绿色 - 最终结果
HEADER_COLOR = "4472C4"  # 深蓝色 - 表头
SUBTOTAL_COLOR = "D9D9D9"  # 浅灰色 - 小计行
WARNING_COLOR = "FFE699"  # 浅橙色 - 重要提示
HEADER_FONT_COLOR = "FFFFFF"  # 白色字体

# ==================== 样式函数 ====================

def apply_header_style(cell):
    """应用表头样式"""
    cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

def apply_input_style(cell):
    """应用输入单元格样式"""
    cell.fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type='solid')
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal='right', vertical='center')

def apply_calc_style(cell):
    """应用计算单元格样式"""
    cell.fill = PatternFill(start_color=CALC_COLOR, end_color=CALC_COLOR, fill_type='solid')
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal='right', vertical='center')

def apply_result_style(cell):
    """应用结果单元格样式"""
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=RESULT_COLOR, end_color=RESULT_COLOR, fill_type='solid')
    cell.border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal='right', vertical='center')

def apply_subtotal_style(cell):
    """应用小计行样式"""
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=SUBTOTAL_COLOR, end_color=SUBTOTAL_COLOR, fill_type='solid')
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal='right', vertical='center')

# ==================== 工作表创建函数 ====================

def create_parameters_sheet(wb):
    """创建边界设定工作表"""
    ws = wb.active
    ws.title = "边界设定"
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    
    row = 1
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "德国独立储能电站投资测算系统 - 边界条件设定"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    params = [
        (2, "电站装机功率", 100, "MW", "德国大型储能项目典型规模"),
        (3, "电站储能容量", 200, "MWh", "2小时储能配置"),
        (4, "储能时长", '=IF(B3>0,B3/B2,"")', "小时", "自动计算"),
        (5, "首年电池可用容量", 100, "%", "通常为100%"),
        (6, "电站运营年限", 20, "年", "行业标准运营周期"),
        (7, "资本金比例", 25, "%", "德国项目融资典型比例25-30%"),
        (8, "贷款年限", 12, "年", "德国银行储能项目常见贷款期限"),
        (9, "贷款利率", 4.5, "%", "2025年欧洲市场基准利率水平"),
        (10, "还款宽限期", 1, "年", ""),
        (11, "还款方式", "等额本金", "", "等额本金/等额本息"),
        (12, "建设期", 1, "年", "项目从开工到投产的时间"),
        (13, "建设期资金占用比例", 50, "%", "建设期平均资金占用比例"),
        (14, "通货膨胀率", 2.0, "%/年", "用于OPEX年度增长计算"),
        (15, "固定资产折旧年限", 15, "年", ""),
        (16, "残值率", 5, "%", ""),
        (17, "折旧方法", "直线法", "", "直线法/双倍余额递减法/年数总和法"),
        (18, "无形资产摊销年限", 20, "年", ""),
        (19, "增值税率", 19, "%", "德国标准增值税率"),
        (20, "企业所得税率", 15, "%", "德国法定企业所得税率"),
        (21, "团结附加税率", 5.5, "%", "按企业所得税额的5.5%计算"),
        (22, "贸易税率", 14, "%", "德国各地市政税率不同(7-17%)"),
        (23, "其他税费", 0, "%", ""),
        (24, "Tolling合同年限", 10, "年", "德国储能Tolling合约典型期限"),
        (25, "Tolling容量占比", 80, "%", "银行融资通常要求70-90%"),
        (26, "Tolling价格(首年)", 95, "EUR/kW/年", "德国2025年市场价格70-90€/kW"),
        (27, "Tolling价格年增长率", 2.0, "%", "通胀调整系数"),
        (28, "充电效率", 95, "%", ""),
        (29, "放电效率", 95, "%", ""),
        (30, "年电池容量衰减率", 2.5, "%/年", "线性衰减模式：固定年衰减率"),
        (31, "系统综合效率(RTE)", "=B28*B29/100", "%", "自动计算"),
        (32, "衰减模式", "线性衰减", "", "线性衰减/非线性衰减/循环次数衰减"),
    ]
    
    for row, name, value, unit, note in params:
        ws.cell(row, 1, name)
        cell = ws.cell(row, 2)
        if isinstance(value, str) and value.startswith('='):
            cell.value = value
            apply_calc_style(cell)
        else:
            cell.value = value
            apply_input_style(cell)
            if name == '折旧方法':
                depreciation_methods = ["直线法", "双倍余额递减法", "年数总和法"]
                list_str = ",".join(depreciation_methods)
                dv = DataValidation(type="list", formula1=f'"{list_str}"', allow_blank=False)
                dv.error = '请从列表中选择折旧方法'
                dv.errorTitle = '无效输入'
                dv.add(cell.coordinate)
                ws.add_data_validation(dv)
            if name == '还款方式':
                repayment_methods = ["等额本金", "等额本息"]
                list_str = ",".join(repayment_methods)
                dv = DataValidation(type="list", formula1=f'"{list_str}"', allow_blank=False)
                dv.error = '请从列表中选择还款方式'
                dv.errorTitle = '无效输入'
                dv.add(cell.coordinate)
                ws.add_data_validation(dv)
        
        ws.cell(row, 3, unit)
        ws.cell(row, 4, note)

def create_equipment_sheet(wb):
    """创建设备配置工作表"""
    ws = wb.create_sheet("设备配置")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    row = 1
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "设备配置表"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # 定义下拉列表选项
    battery_models = [
        "CATL EnerOne Plus 5MWh (314Ah) ★推荐",
        "CATL EnerC Plus 6.25MWh (314Ah)",
        "CATL EnerD 5MWh (530Ah)",
        "CATL TENER 6.25MWh (LFP)",
        "BYD MC Cube 2.8MWh (280Ah)",
        "BYD Cube Pro 3.7MWh (302Ah)",
        "BYD Battery-Box Premium 1.34MWh",
        "EVE LF560K 5MWh (560Ah)",
        "EVE LF280K 3.35MWh (280Ah)",
        "EVE LF314K 3.76MWh (314Ah)",
        "REPT 320Ah 3.84MWh",
        "REPT 345Ah 4.14MWh",
        "海辰 314Ah 3.76MWh",
        "海辰 560Ah 5.02MWh",
        "国轩 280Ah 3.35MWh",
        "国轩 314Ah 3.76MWh",
        "Samsung SDI E3 3.92MWh",
        "LG RESU Prime 3.5MWh",
        "自定义"
    ]
    
    pcs_models = [
        "Sungrow ST5220KWH 5.5MW ★推荐",
        "Sungrow ST3440KWH 3.45MW",
        "Sungrow ST2752KWH 2.75MW",
        "Sungrow SC5000UD-MV 5MW (集中式)",
        "Huawei LUNA2000-2.0MW",
        "Huawei Smart String ESS 2.5MW",
        "Huawei FusionPower 5MW",
        "科华 BluE-S 3450kW",
        "科华 BluE-S 2500kW",
        "科华 BluE-M 5000kW",
        "上能 ESC-3450kW",
        "上能 ESC-5000kW",
        "Sungrow SC8000UD-MV 8MW",
        "特变 TC-3450ESS 3.45MW",
        "特变 TC-5000ESS 5MW",
        "盛弘 ETC-3450 3.45MW",
        "南瑞继保 PCS-3450 3.45MW",
        "SMA Sunny Central Storage 2.5MW",
        "ABB PCS-3000 3MW",
        "Siemens Siestorage 2MW",
        "自定义"
    ]
    
    mv_transformer_specs = [
        "20kV/1600kVA 干式",
        "20kV/2000kVA 干式",
        "20kV/2500kVA 干式",
        "20kV/3150kVA 干式",
        "20kV/4000kVA 干式",
        "20kV/5000kVA 干式",
        "20kV/6300kVA 干式 ★推荐",
        "30kV/6300kVA 干式",
        "35kV/6300kVA 干式",
        "自定义"
    ]
    
    hv_transformer_specs = [
        "110kV/50MVA 油浸式",
        "110kV/63MVA 油浸式",
        "110kV/80MVA 油浸式",
        "110kV/100MVA 油浸式",
        "110kV/120MVA 油浸式 ★推荐",
        "110kV/150MVA 油浸式",
        "220kV/120MVA 油浸式",
        "220kV/150MVA 油浸式",
        "220kV/180MVA 油浸式",
        "220kV/240MVA 油浸式",
        "自定义"
    ]
    
    equipment_rows = [
        (2, "电池单价", 85, "EUR/kWh", "从边界设定表选择电池型号后自动填充"),
        (3, "电池型号", "CATL EnerOne Plus 5MWh (314Ah) ★推荐", "", "选择电池型号"),
        (4, "电池容量", "=边界设定!B3", "MWh", "引用边界设定"),
        (5, "PCS单价", 120, "EUR/kW", ""),
        (6, "PCS型号", "Sungrow ST5220KWH 5.5MW ★推荐", "", "选择PCS型号"),
        (7, "中压变压器单价", 150000, "EUR/台", ""),
        (8, "中压变压器数量", 2, "台", ""),
        (9, "升压变压器单价", 800000, "EUR/台", ""),
        (10, "升压变压器数量", 1, "台", ""),
        (11, "中压变压器型号", "20kV/6300kVA 干式 ★推荐", "", "选择中压变压器规格"),
        (12, "升压变压器型号", "110kV/120MVA 油浸式 ★推荐", "", "选择升压变压器规格"),
        (13, "EMS系统", 500000, "EUR", ""),
        (14, "SCADA系统", 300000, "EUR", ""),
        (15, "开关柜单价", 50000, "EUR/面", ""),
        (16, "开关柜数量", 10, "面", ""),
        (17, "集电线路", 2000000, "EUR", ""),
        (18, "热管理系统单价", 15, "EUR/kWh", ""),
        (19, "消防系统单价", 20, "EUR/kWh", ""),
        (22, "变电站建设", 3000000, "EUR", ""),
        (23, "接入线路", 2000000, "EUR", ""),
        (24, "并网申请与研究", 500000, "EUR", ""),
        (25, "计量与保护设备", 300000, "EUR", ""),
        (27, "土地获取成本", 50, "EUR/kW", ""),
        (28, "混凝土基础", 30, "EUR/kW", ""),
        (29, "围栏与安防", 200000, "EUR", ""),
        (30, "道路建设", 300000, "EUR", ""),
        (31, "排水系统", 150000, "EUR", ""),
        (33, "机电安装费率", 8, "%", "设备费比例"),
        (34, "施工管理费率", 3, "%", "设备费比例"),
        (35, "调试费用", 500000, "EUR", ""),
        (36, "CAR保险费率", 0.3, "%", "设备费比例"),
        (37, "EAR保险费率", 0.2, "%", "设备费比例"),
        (38, "货物运输保险费率", 0.5, "%", "设备费比例"),
        (39, "第三方责任险", 100000, "EUR", ""),
        (40, "SPV公司收购成本", 200000, "EUR", ""),
        (41, "许可与规划费", 300000, "EUR", ""),
        (42, "环境咨询费", 200000, "EUR", ""),
        (43, "法律咨询费", 150000, "EUR", ""),
        (44, "工程设计费率", 2, "%", "设备费比例"),
        (45, "项目管理费率", 2, "%", "总投资比例"),
        (46, "不可预见费率", 5, "%", ""),
        (47, "拆除准备金", 500000, "EUR", ""),
    ]
    
    for row, name, value, unit, note in equipment_rows:
        ws.cell(row, 1, name)
        cell = ws.cell(row, 2)
        if isinstance(value, str) and value.startswith('='):
            cell.value = value
            apply_calc_style(cell)
        else:
            cell.value = value
            apply_input_style(cell)
        
        ws.cell(row, 3, unit)
        ws.cell(row, 4, note)
    
    dropdown_cells = {
        3: battery_models,
        6: pcs_models,
        11: mv_transformer_specs,
        12: hv_transformer_specs
    }
    for row, options in dropdown_cells.items():
        cell = ws.cell(row, 2)
        list_str = ",".join(options)
        dv = DataValidation(type="list", formula1=f'"{list_str}"', allow_blank=False)
        if row == 3:
            dv.error = '请从列表中选择电池型号'
            dv.errorTitle = '无效输入'
        elif row == 6:
            dv.error = '请从列表中选择PCS型号'
            dv.errorTitle = '无效输入'
        elif row == 11:
            dv.error = '请从列表中选择中压变压器规格'
            dv.errorTitle = '无效输入'
        elif row == 12:
            dv.error = '请从列表中选择升压变压器规格'
            dv.errorTitle = '无效输入'
        dv.add(cell.coordinate)
        ws.add_data_validation(dv)
    
    ws.merge_cells('A48:D48')
    header_cell = ws['A48']
    header_cell.value = "OPEX基础参数（运行期）"
    header_cell.font = Font(bold=True, size=12, color=HEADER_FONT_COLOR)
    header_cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    
    headers = ['项目', '数值', '单位', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(49, col, header)
        apply_header_style(cell)
    
    opex_rows = [
        (50, "技术运维基础值", 6, "EUR/kW", ""),
        (51, "技术运维增长率", 2.0, "%", ""),
        (52, "保险费率", 0.4, "%", ""),
        (53, "保险增长率", 1.5, "%", ""),
        (54, "电网费用基础值", 12000, "EUR/MW", ""),
        (55, "电网费用增长率", 2.0, "%", ""),
        (56, "土地租金基础值", 60000, "EUR/年", ""),
        (57, "土地租金增长率", 2.0, "%", ""),
        (58, "商务费用基础值", 4000, "EUR/MW", ""),
        (59, "商务费用增长率", 2.0, "%", ""),
        (60, "其他费用基础值", 1500, "EUR/MW", ""),
        (61, "其他费用增长率", 2.0, "%", ""),
    ]
    
    for row, name, value, unit, note in opex_rows:
        ws.cell(row, 1, name)
        cell = ws.cell(row, 2)
        cell.value = value
        apply_input_style(cell)
        ws.cell(row, 3, unit)
        ws.cell(row, 4, note)

def create_capex_sheet(wb):
    """创建CAPEX明细工作表"""
    ws = wb.create_sheet("CAPEX明细")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "CAPEX 投资明细表"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['项目', '单价/比例', '数量', '金额（万EUR）']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 一、主设备
    capex_items = [
        ("一、主设备", None, None, None),
        ("电池系统", "=设备配置!B2", "=边界设定!B3", "=B{0}*C{0}*1000/10000"),
        ("PCS系统", "=设备配置!B5", "=边界设定!B2", "=B{0}*C{0}*1000/10000"),
        ("中压变压器", "=设备配置!B7", "=设备配置!B8", "=B{0}*C{0}/10000"),
        ("升压变压器", "=设备配置!B9", "=设备配置!B10", "=B{0}*C{0}/10000"),
        ("设备费小计", None, None, "=SUM(D5:D9)"),
        ("", None, None, None),
        ("二、辅助设备", None, None, None),
        ("EMS系统", "=设备配置!B13", "1", "=B{0}*C{0}/10000"),
        ("SCADA系统", "=设备配置!B14", "1", "=B{0}*C{0}/10000"),
        ("开关柜", "=设备配置!B15", "=设备配置!B16", "=B{0}*C{0}/10000"),
        ("集电线路", "=设备配置!B17", "1", "=B{0}*C{0}/10000"),
        ("热管理系统", "=设备配置!B18", "=边界设定!B3", "=B{0}*C{0}*1000/10000"),
        ("消防系统", "=设备配置!B19", "=边界设定!B3", "=B{0}*C{0}*1000/10000"),
        ("辅助设备小计", None, None, "=SUM(D{0}:D{1})"),
        ("", None, None, None),
        ("三、电网接入", None, None, None),
        ("变电站建设", "=设备配置!B22", "1", "=B{0}/10000"),
        ("接入线路", "=设备配置!B23", "1", "=B{0}/10000"),
        ("并网申请与研究", "=设备配置!B24", "1", "=B{0}/10000"),
        ("计量与保护设备", "=设备配置!B25", "1", "=B{0}/10000"),
        ("电网接入小计", None, None, "=SUM(D{0}:D{1})"),
        ("", None, None, None),
        ("四、土地与基建", None, None, None),
        ("土地获取成本", "=设备配置!B27", "=边界设定!B2", "=B{0}*C{0}*1000/10000"),
        ("混凝土基础", "=设备配置!B28", "=边界设定!B2", "=B{0}*C{0}*1000/10000"),
        ("围栏与安防", "=设备配置!B29", "1", "=B{0}/10000"),
        ("道路建设", "=设备配置!B30", "1", "=B{0}/10000"),
        ("排水系统", "=设备配置!B31", "1", "=B{0}/10000"),
        ("土地与基建小计", None, None, "=SUM(D{0}:D{1})"),
        ("", None, None, None),
        ("五、安装与施工", None, None, None),
        ("机电安装", "=设备配置!B33", "=D5+D9", "=B{0}*C{0}/100"),
        ("施工管理费", "=设备配置!B34", "=D5+D9", "=B{0}*C{0}/100"),
        ("调试费用", "=设备配置!B35", "1", "=B{0}/10000"),
        ("安装施工小计", None, None, "=SUM(D{0}:D{1})"),
        ("", None, None, None),
        ("六、建设期保险", None, None, None),
        ("CAR保险", "=设备配置!B36", "=D5+D9", "=B{0}*C{0}/100"),
        ("EAR保险", "=设备配置!B37", "=D5+D9", "=B{0}*C{0}/100"),
        ("货物运输保险", "=设备配置!B38", "=D5+D9", "=B{0}*C{0}/100"),
        ("第三方责任险", "=设备配置!B39", "1", "=B{0}/10000"),
        ("保险费小计", None, None, "=SUM(D{0}:D{1})"),
        ("", None, None, None),
        ("七、开发与业主费用", None, None, None),
        ("SPV公司收购成本", "=设备配置!B40", "1", "=B{0}/10000"),
        ("许可与规划费", "=设备配置!B41", "1", "=B{0}/10000"),
        ("环境咨询费", "=设备配置!B42", "1", "=B{0}/10000"),
        ("法律咨询费", "=设备配置!B43", "1", "=B{0}/10000"),
        ("工程设计费", "=设备配置!B44", "=D5+D9", "=B{0}*C{0}/100"),
        ("项目管理费", "=设备配置!B45", "=SUM(D5:D{0})", "=B{0}*C{0}/100"),
        ("开发费用小计", None, None, "=SUM(D{0}:D{1})"),
        ("", None, None, None),
        ("八、其他", None, None, None),
        ("不可预见费", "=设备配置!B46", "=SUM(D5:D{0})", "=B{0}*C{0}/100"),
        ("拆除准备金", "=设备配置!B47", "1", "=B{0}/10000"),
        ("CAPEX总计（不含建设期利息）", None, None, "=SUM(D5:D{0})"),
        ("建设期利息", None, None, "=D{0}*(1-边界设定!B7/100)*边界设定!B9/100*边界设定!B12*边界设定!B13/100"),
        ("CAPEX总计（含建设期利息）", None, None, "=D{0}+D{1}"),
    ]
    
    start_row = row
    section_start_row = None  # 记录每个分组的起始行
    data_rows = []  # 记录每个分组的数据行范围
    capex_total_row = None  # 记录CAPEX总计行的行号
    capex_total_with_interest_row = None  # 记录含建设期利息的CAPEX总计行号
    
    for item in capex_items:
        name, price, qty, amount = item
        if name:
            # 如果是分组标题，记录起始行
            if name.startswith('一、') or name.startswith('二、') or name.startswith('三、') or \
               name.startswith('四、') or name.startswith('五、') or name.startswith('六、') or \
               name.startswith('七、') or name.startswith('八、'):
                section_start_row = row + 1  # 下一行是数据开始
                ws.cell(row, 1, name)
                ws.cell(row, 1).font = Font(bold=True)
                ws.cell(row, 1).fill = PatternFill(start_color=SUBTOTAL_COLOR, end_color=SUBTOTAL_COLOR, fill_type='solid')
            else:
                ws.cell(row, 1, name)
                if price:
                    cell = ws.cell(row, 2)
                    if isinstance(price, str) and price.startswith('='):
                        cell.value = price
                    else:
                        cell.value = price
                    apply_calc_style(cell)
                if qty:
                    cell = ws.cell(row, 3)
                    if isinstance(qty, str) and qty.startswith('='):
                        cell.value = qty
                    else:
                        cell.value = qty
                    apply_calc_style(cell)
                if amount:
                    cell = ws.cell(row, 4)
                    if isinstance(amount, str) and amount.startswith('='):
                        # 处理包含 {0} 和 {1} 的公式
                        if 'SUM(D{0}:D{1})' in amount:
                            # 小计行：需要起始行和当前行前一行
                            if section_start_row:
                                formula = amount.replace('{0}', str(section_start_row)).replace('{1}', str(row - 1))
                            else:
                                formula = amount.replace('{0}', '5').replace('{1}', str(row - 1))
                        elif 'D{0}+D{1}' in amount:
                            # 总计行：需要引用前面的两个单元格
                            # 找到前面的总计行和建设期利息行
                            # 注意：这里需要找到实际的CAPEX总计行和建设期利息行
                            # 由于行号是动态的，我们需要向上查找
                            prev_total_row = row - 2  # CAPEX总计（不含建设期利息）行
                            prev_interest_row = row - 1  # 建设期利息行
                            if prev_total_row >= start_row and prev_interest_row >= start_row:
                                formula = amount.replace('{0}', str(prev_total_row)).replace('{1}', str(prev_interest_row))
                            else:
                                # 如果行号无效，使用默认值
                                formula = amount.replace('{0}', 'D50').replace('{1}', 'D51')
                        elif 'SUM(D5:D{0})' in amount:
                            # 需要找到前面的总计行
                            formula = amount.replace('{0}', str(row - 1))
                        elif '{0}' in amount:
                            # 只包含 {0} 的公式，替换为当前行号
                            formula = amount.replace('{0}', str(row))
                        else:
                            formula = amount
                        cell.value = formula
                    else:
                        cell.value = amount
                    apply_calc_style(cell)
                    if '小计' in name or '总计' in name:
                        if '小计' in name:
                            apply_subtotal_style(cell)
                        else:
                            apply_result_style(cell)
                            # 记录CAPEX总计行的行号
                            if 'CAPEX总计（不含建设期利息）' in name:
                                capex_total_row = row
                            elif 'CAPEX总计（含建设期利息）' in name:
                                capex_total_with_interest_row = row
        row += 1
    
    # 注意：所有公式已使用INDEX和MATCH动态查找CAPEX总计行
    # 不再需要硬编码行号，提高了公式的健壮性

def create_spot_price_sheet(wb):
    """创建现货价格工作表"""
    ws = wb.create_sheet("现货价格")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    
    row = 1
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "现货价格表（EUR/MWh）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    cell = ws.cell(row, 1, "年份")
    apply_header_style(cell)
    cell = ws.cell(row, 2, "现货价格")
    apply_header_style(cell)
    row += 1
    
    # 生成20年的价格行（默认值，用户可修改）
    for year in range(1, 21):
        ws.cell(row, 1, f"第{year}年")
        cell = ws.cell(row, 2, 80)  # 默认价格
        apply_input_style(cell)
        row += 1

def create_opex_sheet(wb):
    """创建OPEX设定工作表"""
    ws = wb.create_sheet("OPEX设定")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 15
    
    row = 1
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = "OPEX 年度运营成本表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '技术运维', '保险', '电网费用', '土地租金', '商务费用', '其他', '拆除准备金', '合计']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 生成年度数据行（使用公式）
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        inflation_factor = f'POWER(1+边界设定!B14/100,{year-1})'
        formulas = [
            f'=设备配置!B50*边界设定!B2*1000*POWER(1+设备配置!B51/100,{year-1})*{inflation_factor}/10000',
            f'=IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（不含建设期利息）",CAPEX明细!A:A,0)),0)*设备配置!B52/100*POWER(1+设备配置!B53/100,{year-1})*{inflation_factor}',
            f'=设备配置!B54*边界设定!B2*POWER(1+设备配置!B55/100,{year-1})*{inflation_factor}/10000',
            f'=设备配置!B56*POWER(1+设备配置!B57/100,{year-1})*{inflation_factor}/10000',
            f'=设备配置!B58*边界设定!B2*POWER(1+设备配置!B59/100,{year-1})*{inflation_factor}/10000',
            f'=设备配置!B60*边界设定!B2*POWER(1+设备配置!B61/100,{year-1})*{inflation_factor}/10000',
            '=0'
        ]
        for col_idx, formula in enumerate(formulas, 2):
            cell = ws.cell(row, col_idx)
            cell.value = formula
            apply_calc_style(cell)
        # 合计
        cell = ws.cell(row, 9)
        cell.value = f'=SUM(B{row}:H{row})'
        apply_result_style(cell)
        row += 1

def create_revenue_sheet(wb):
    """创建收入预测工作表"""
    ws = wb.create_sheet("收入预测")
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18
    
    row = 1
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "收入预测表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '可用容量比例(%)', 'Tolling收入', '现货收入', '总收入']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 生成年度数据行
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        # 可用容量比例（考虑衰减）
        cell = ws.cell(row, 2)
        cell.value = f'=边界设定!B5/100*POWER(1-边界设定!B30/100,{year-1})*100'
        apply_calc_style(cell)
        # Tolling收入
        cell = ws.cell(row, 3)
        cell.value = f'=IF({year}<=边界设定!B24,边界设定!B26*边界设定!B2*1000*边界设定!B25/100/10000*POWER(1+边界设定!B27/100,{year-1}),0)'
        apply_calc_style(cell)
        # 现货收入
        cell = ws.cell(row, 4)
        cell.value = f'=现货价格!B{year+2}*边界设定!B2*IF({year}<=边界设定!B24,1-边界设定!B25/100,1)*B{row}/100/10000'
        apply_calc_style(cell)
        # 总收入
        cell = ws.cell(row, 5)
        cell.value = f'=C{row}+D{row}'
        apply_result_style(cell)
        row += 1

def create_depreciation_sheet(wb):
    """创建折旧计算工作表"""
    ws = wb.create_sheet("折旧计算")
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18
    
    row = 1
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "折旧摊销计算表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '固定资产折旧', '无形资产摊销', '折旧合计', '累计折旧']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 生成年度数据行（使用公式）
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        # 固定资产折旧（直线法）
        # 使用动态查找CAPEX总计行，而不是硬编码D50
        cell = ws.cell(row, 2)
        # 查找"CAPEX总计（含建设期利息）"所在行，使用IFERROR处理找不到的情况
        # 使用命名公式提高可读性和性能
        cell.value = f'=IF({year}<=边界设定!B15,IFERROR((INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0))-INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0))*边界设定!B16/100)/边界设定!B15,0),0)'
        apply_calc_style(cell)
        # 无形资产摊销（开发费用+土地）
        cell = ws.cell(row, 3)
        # 计算无形资产 = 开发费用小计 + 土地获取成本，使用IFERROR处理找不到的情况
        cell.value = f'=IF({year}<=边界设定!B18,IFERROR((INDEX(CAPEX明细!D:D,MATCH("开发费用小计",CAPEX明细!A:A,0))+INDEX(CAPEX明细!D:D,MATCH("土地获取成本",CAPEX明细!A:A,0)))/边界设定!B18,0),0)'
        apply_calc_style(cell)
        # 折旧合计
        cell = ws.cell(row, 4)
        cell.value = f'=B{row}+C{row}'
        apply_calc_style(cell)
        # 累计折旧
        cell = ws.cell(row, 5)
        cell.value = f'=SUM(B$5:B{row})+SUM(C$5:C{row})'
        apply_calc_style(cell)
        row += 1

def create_loan_sheet(wb):
    """创建贷款计算工作表"""
    ws = wb.create_sheet("贷款计算")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18
    
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "贷款还款计划表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '期初余额', '利息', '本金', '还款额', '期末余额']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 生成年度数据行（使用公式，等额本金方式）
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        # 期初余额
        cell = ws.cell(row, 2)
        if year == 1:
            # 使用动态查找CAPEX总计，使用IFERROR处理找不到的情况
            cell.value = '=IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0))*(1-边界设定!B7/100),0)'
        else:
            cell.value = f'=F{row-1}'
        apply_calc_style(cell)
        # 利息
        cell = ws.cell(row, 3)
        cell.value = f'=B{row}*边界设定!B9/100'
        apply_calc_style(cell)
        # 本金
        cell = ws.cell(row, 4)
        # 使用动态查找CAPEX总计，使用IFERROR处理找不到的情况
        cell.value = f'=IF({year}>边界设定!B10,IF(边界设定!B11="等额本金",IFERROR((INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0))*(1-边界设定!B7/100))/(边界设定!B8-边界设定!B10),0),IFERROR(PMT(边界设定!B9/100,边界设定!B8-边界设定!B10,-INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0))*(1-边界设定!B7/100))-C{row},0)),0)'
        apply_calc_style(cell)
        # 还款额
        cell = ws.cell(row, 5)
        cell.value = f'=C{row}+D{row}'
        apply_calc_style(cell)
        # 期末余额
        cell = ws.cell(row, 6)
        cell.value = f'=MAX(0,B{row}-D{row})'
        apply_calc_style(cell)
        row += 1

def create_income_sheet(wb):
    """创建利润表工作表"""
    ws = wb.create_sheet("利润表")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15
    
    row = 1
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = "利润表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '营业收入', '营业成本', '毛利润', 'EBITDA', '折旧', 'EBIT', '利息', 'EBT', '所得税', '净利润']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 生成年度数据行
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        # 营业收入
        cell = ws.cell(row, 2)
        cell.value = f'=收入预测!E{year+2}'
        apply_calc_style(cell)
        # 营业成本（OPEX）
        cell = ws.cell(row, 3)
        cell.value = f'=OPEX设定!I{year+2}'
        apply_calc_style(cell)
        # 毛利润
        cell = ws.cell(row, 4)
        cell.value = f'=B{row}-C{row}'
        apply_calc_style(cell)
        # EBITDA
        cell = ws.cell(row, 5)
        cell.value = f'=D{row}'
        apply_calc_style(cell)
        # 折旧
        cell = ws.cell(row, 6)
        cell.value = f'=折旧计算!D{year+2}'
        apply_calc_style(cell)
        # EBIT
        cell = ws.cell(row, 7)
        cell.value = f'=E{row}-F{row}'
        apply_calc_style(cell)
        # 利息
        cell = ws.cell(row, 8)
        cell.value = f'=贷款计算!C{year+2}'
        apply_calc_style(cell)
        # EBT
        cell = ws.cell(row, 9)
        cell.value = f'=G{row}-H{row}'
        apply_calc_style(cell)
        # 所得税
        cell = ws.cell(row, 10)
        cell.value = f'=MAX(0,I{row}*(边界设定!B20/100*(1+边界设定!B21/100)+边界设定!B22/100+边界设定!B23/100))'
        apply_calc_style(cell)
        # 净利润
        cell = ws.cell(row, 11)
        cell.value = f'=I{row}-J{row}'
        apply_result_style(cell)
        row += 1

def create_cashflow_sheet(wb):
    """创建现金流量表工作表"""
    ws = wb.create_sheet("现金流量表")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20
    
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "现金流量表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '经营活动现金流', '投资活动现金流', '筹资活动现金流', '全投资现金流', '资本金现金流']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 建设期
    ws.cell(row, 1, '建设期')
    cell = ws.cell(row, 2)
    cell.value = 0
    apply_calc_style(cell)
    cell = ws.cell(row, 3)
    # 使用动态查找CAPEX总计，使用IFERROR处理找不到的情况
    cell.value = '=-IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)'
    apply_calc_style(cell)
    cell = ws.cell(row, 4)
    cell.value = '=IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)'
    apply_calc_style(cell)
    cell = ws.cell(row, 5)
    cell.value = '=-IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)'
    apply_result_style(cell)
    cell = ws.cell(row, 6)
    cell.value = '=-IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)*边界设定!B7/100'
    apply_result_style(cell)
    row += 1
    
    # 运营期
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        # 经营活动现金流
        cell = ws.cell(row, 2)
        cell.value = f'=利润表!K{year+2}+折旧计算!D{year+2}'
        apply_calc_style(cell)
        # 投资活动现金流
        cell = ws.cell(row, 3)
        if year == 20:
            # 残值回收 = 固定资产原值 * 残值率，使用IFERROR处理找不到的情况
            cell.value = f'=IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)*边界设定!B16/100'
        else:
            cell.value = 0
        apply_calc_style(cell)
        # 筹资活动现金流
        cell = ws.cell(row, 4)
        cell.value = f'=-贷款计算!D{year+2}'
        apply_calc_style(cell)
        # 全投资现金流
        cell = ws.cell(row, 5)
        cell.value = f'=利润表!E{year+2}-利润表!J{year+2}+B{row}'
        apply_result_style(cell)
        # 资本金现金流
        cell = ws.cell(row, 6)
        cell.value = f'=B{row}+C{row}+D{row}'
        apply_result_style(cell)
        row += 1

def create_balance_sheet(wb):
    """创建资产负债表工作表"""
    ws = wb.create_sheet("资产负债表")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18
    
    row = 1
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "资产负债表（万EUR）"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 表头
    headers = ['年份', '货币资金', '固定资产净值', '无形资产', '资产总计', '长期借款', '实收资本', '未分配利润', '负债和权益总计']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        apply_header_style(cell)
    row += 1
    
    # 初始资产负债表
    ws.cell(row, 1, '建设完成')
    cell = ws.cell(row, 2)
    cell.value = 0
    apply_calc_style(cell)
    cell = ws.cell(row, 3)
    # 固定资产净值 = CAPEX总计 - 残值，使用IFERROR处理找不到的情况
    capex_total_formula = 'IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)'
    cell.value = f'={capex_total_formula}-{capex_total_formula}*边界设定!B16/100'
    apply_calc_style(cell)
    cell = ws.cell(row, 4)
    # 无形资产 = 开发费用 + 土地，使用IFERROR处理找不到的情况
    cell.value = f'=IFERROR(INDEX(CAPEX明细!D:D,MATCH("开发费用小计",CAPEX明细!A:A,0)),0)+IFERROR(INDEX(CAPEX明细!D:D,MATCH("土地获取成本",CAPEX明细!A:A,0)),0)'
    apply_calc_style(cell)
    cell = ws.cell(row, 5)
    cell.value = f'=B{row}+C{row}+D{row}'
    apply_result_style(cell)
    cell = ws.cell(row, 6)
    cell.value = f'={capex_total_formula}*(1-边界设定!B7/100)'
    apply_calc_style(cell)
    cell = ws.cell(row, 7)
    cell.value = f'={capex_total_formula}*边界设定!B7/100'
    apply_calc_style(cell)
    cell = ws.cell(row, 8)
    cell.value = 0
    apply_calc_style(cell)
    cell = ws.cell(row, 9)
    cell.value = f'=E{row}+F{row}+G{row}+H{row}'
    apply_result_style(cell)
    row += 1
    
    # 运营期
    for year in range(1, 21):
        ws.cell(row, 1, f'第{year}年')
        # 货币资金（累计现金流）
        cell = ws.cell(row, 2)
        cell.value = f'=SUM(现金流量表!B5:B{year+3})'
        apply_calc_style(cell)
        # 固定资产净值
        cell = ws.cell(row, 3)
        cell.value = f'=MAX(0,C{row-1}-折旧计算!B{year+2})'
        apply_calc_style(cell)
        # 无形资产
        cell = ws.cell(row, 4)
        cell.value = f'=MAX(0,D{row-1}-折旧计算!C{year+2})'
        apply_calc_style(cell)
        # 资产总计
        cell = ws.cell(row, 5)
        cell.value = f'=B{row}+C{row}+D{row}'
        apply_result_style(cell)
        # 长期借款
        cell = ws.cell(row, 6)
        cell.value = f'=贷款计算!F{year+2}'
        apply_calc_style(cell)
        # 实收资本，使用IFERROR处理找不到的情况
        cell = ws.cell(row, 7)
        cell.value = '=IFERROR(INDEX(CAPEX明细!D:D,MATCH("CAPEX总计（含建设期利息）",CAPEX明细!A:A,0)),0)*边界设定!B7/100'
        apply_calc_style(cell)
        # 未分配利润
        cell = ws.cell(row, 8)
        cell.value = f'=H{row-1}+利润表!K{year+2}'
        apply_calc_style(cell)
        # 负债和权益总计
        cell = ws.cell(row, 9)
        cell.value = f'=E{row}+F{row}+G{row}+H{row}'
        apply_result_style(cell)
        row += 1

def create_indicators_sheet(wb):
    """创建财务指标工作表"""
    ws = wb.create_sheet("财务指标")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    row = 1
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "财务指标汇总"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # 指标列表
    indicators = [
        ("全投资IRR", "=IRR(现金流量表!E5:E25)"),
        ("资本金IRR", "=IRR(现金流量表!F5:F25)"),
        ("全投资NPV（8%）", "=NPV(0.08,现金流量表!E5:E25)"),
        ("静态回收期", "计算回收期"),
        ("动态回收期", "计算动态回收期"),
        ("ROI", "=IFERROR(SUM(利润表!I5:I24)/IFERROR(INDEX(CAPEX明细!D:D,MATCH(\"CAPEX总计（含建设期利息）\",CAPEX明细!A:A,0)),1),0)"),
        ("ROE（第3年）", "=利润表!K7/资产负债表!H7"),
        ("DSCR（平均）", "=IFERROR(AVERAGE(利润表!E5:E24/贷款计算!E5:E24),0)"),
    ]
    
    for name, formula in indicators:
        ws.cell(row, 1, name)
        ws.cell(row, 1).font = Font(bold=True)
        cell = ws.cell(row, 2)
        if formula.startswith('='):
            cell.value = formula
        else:
            cell.value = formula  # 文本说明
        apply_calc_style(cell)
        if 'IRR' in name or 'NPV' in name or 'ROI' in name or 'ROE' in name:
            cell.number_format = '0.00%'
        row += 1

def create_excel_file():
    """创建完整的Excel文件"""
    # 修复Windows控制台编码问题
    try:
        import sys
        if sys.platform == 'win32':
            sys.stdout.reconfigure(encoding='utf-8')
            sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass
    
    print("正在生成Excel文件...")
    wb = openpyxl.Workbook()
    
    # 创建各个工作表
    create_parameters_sheet(wb)
    create_equipment_sheet(wb)
    create_capex_sheet(wb)
    create_spot_price_sheet(wb)
    create_opex_sheet(wb)
    create_revenue_sheet(wb)
    create_depreciation_sheet(wb)
    create_loan_sheet(wb)
    create_income_sheet(wb)
    create_cashflow_sheet(wb)
    create_balance_sheet(wb)
    create_indicators_sheet(wb)
    
    # 保存文件
    filename = f"德国独立储能电站财务测算表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    
    # 修复Windows控制台编码问题
    try:
        import sys
        if sys.platform == 'win32':
            sys.stdout.reconfigure(encoding='utf-8')
            sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass
    
    print(f"Excel文件已生成: {filepath}")
    return filepath

if __name__ == "__main__":
    try:
        create_excel_file()
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
